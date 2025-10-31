"""
Bank File Export Service
Generate CSV/TXT files for bank transfers
"""
import csv
from io import StringIO
from datetime import datetime
from decimal import Decimal


class BankFileExporter:
    """Export payslips to bank transfer file format"""
    
    def __init__(self, payroll_run):
        self.payroll_run = payroll_run
        self.payslips = payroll_run.payslips.filter(
            salary_deposited=False
        ).select_related('employee')
    
    def generate_csv(self):
        """Generate CSV format for bank transfer"""
        buffer = StringIO()
        writer = csv.writer(buffer)
        
        # Header row
        writer.writerow([
            'Employee No',
            'Employee Name',
            'Bank Name',
            'Account Number',
            'Amount',
            'Reference'
        ])
        
        # Data rows
        for payslip in self.payslips:
            emp = payslip.employee
            writer.writerow([
                emp.employee_no,
                f"{emp.last_name}, {emp.first_name}",
                emp.bank_name or 'N/A',
                emp.bank_account or 'N/A',
                f"{payslip.net_pay:.2f}",
                f"Payroll {self.payroll_run.period_start} to {self.payroll_run.period_end}"
            ])
        
        buffer.seek(0)
        return buffer.getvalue()
    
    def generate_bdo_format(self):
        """Generate BDO bank-specific format"""
        buffer = StringIO()
        
        # BDO format: Account Number|Amount|Reference
        for payslip in self.payslips:
            emp = payslip.employee
            if emp.bank_name and 'BDO' in emp.bank_name.upper():
                line = f"{emp.bank_account}|{payslip.net_pay:.2f}|SALARY-{emp.employee_no}\n"
                buffer.write(line)
        
        buffer.seek(0)
        return buffer.getvalue()
    
    def generate_bpi_format(self):
        """Generate BPI bank-specific format"""
        buffer = StringIO()
        
        # BPI format: Account Number,Amount,Reference
        for payslip in self.payslips:
            emp = payslip.employee
            if emp.bank_name and 'BPI' in emp.bank_name.upper():
                line = f"{emp.bank_account},{payslip.net_pay:.2f},SALARY-{emp.employee_no}\n"
                buffer.write(line)
        
        buffer.seek(0)
        return buffer.getvalue()
    
    def generate_summary(self):
        """Generate summary report"""
        total_employees = self.payslips.count()
        total_amount = sum(p.net_pay for p in self.payslips)
        
        summary = f"""
BANK TRANSFER SUMMARY
Payroll Period: {self.payroll_run.period_start} to {self.payroll_run.period_end}
Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

Total Employees: {total_employees}
Total Amount: ₱{total_amount:,.2f}

Breakdown by Bank:
"""
        
        # Group by bank
        banks = {}
        for payslip in self.payslips:
            bank = payslip.employee.bank_name or 'No Bank'
            if bank not in banks:
                banks[bank] = {'count': 0, 'amount': Decimal('0.00')}
            banks[bank]['count'] += 1
            banks[bank]['amount'] += payslip.net_pay
        
        for bank, data in sorted(banks.items()):
            summary += f"  {bank}: {data['count']} employees, ₱{data['amount']:,.2f}\n"
        
        return summary
    
    def mark_as_generated(self):
        """Mark payslips as bank file generated"""
        self.payslips.update(bank_file_generated=True)


def export_to_excel(payroll_run):
    """Export payroll to Excel using openpyxl"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Payroll"
    
    # Headers
    headers = [
        'Employee No', 'Name', 'Department', 'Position',
        'Gross Pay', 'Overtime', 'Total Earnings',
        'SSS', 'PhilHealth', 'Pag-IBIG', 'Tax',
        'Loan Deductions', 'Other Deductions', 'Total Deductions',
        'Net Pay'
    ]
    
    ws.append(headers)
    
    # Style header row
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Data rows
    payslips = payroll_run.payslips.all().select_related('employee')
    
    for payslip in payslips:
        emp = payslip.employee
        ws.append([
            emp.employee_no,
            f"{emp.last_name}, {emp.first_name}",
            emp.department,
            emp.position,
            float(payslip.gross_pay),
            float(payslip.overtime_pay),
            float(payslip.total_earnings),
            float(payslip.sss),
            float(payslip.philhealth),
            float(payslip.pagibig),
            float(payslip.tax),
            float(payslip.loan_deductions),
            float(payslip.other_deductions),
            float(payslip.total_deductions),
            float(payslip.net_pay),
        ])
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer
